#!/usr/bin/env python3
"""
Excel → SQLite 导入脚本 (v3)

将手工整理的 Excel 数据库全量导入 SQLite，作为基线去重依据。

用法:
    python scripts/import_excel.py                    # 默认路径导入
    python scripts/import_excel.py --excel 其他.xlsx   # 自定义路径
    python scripts/import_excel.py --dry-run           # 预览不写入
"""
import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

# 允许从项目根目录或 scripts/ 目录运行
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from loguru import logger

from config.taxonomy import TRACK_TAGS
from models.schema import Deal
from storage import db

# 默认 Excel 路径
DEFAULT_EXCEL = r"C:\Users\陈浩\Desktop\AI创业资讯_更新6.25_商业航天拆分_宽口径tag.xlsx"

# 数据 sheet 名 = track 名（Tag体系 跳过）
DATA_SHEETS = list(TRACK_TAGS.keys())  # ["AI2C", "AI2B", "具身", "ai4S", "前沿科技", "商业航天"]


def _has_chinese(s: str) -> bool:
    """判断字符串是否包含中文字符。"""
    return bool(re.search(r'[一-鿿]', s))


def _clean_val(val) -> str:
    """清洗单元格值：None → ""，datetime → "YYYY-MM-DD"，去首尾空格。"""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    return "" if s.lower() in ("nan", "none") else s


def _import_sheet(ws, track: str) -> list[Deal]:
    """读取单个 sheet，返回 Deal 列表。"""
    deals = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 10:
            continue

        project_name = _clean_val(row[0])
        if not project_name:
            continue  # 空行跳过

        sub_tag = _clean_val(row[1])
        founded_year_raw = _clean_val(row[2])
        title = _clean_val(row[3])
        team = _clean_val(row[4])
        round_ = _clean_val(row[5])
        amount = _clean_val(row[6])
        investors = _clean_val(row[7])
        region = _clean_val(row[8])
        detail = _clean_val(row[9])

        # "未披露" 统一处理
        if founded_year_raw in ("未披露", ""):
            founded_year_raw = ""
        if amount in ("", "未披露", None):
            amount = "未披露"

        # region → region_class
        if not region:
            region_class = "未知"
        elif _has_chinese(region):
            region_class = "国内"
        else:
            region_class = "海外"

        try:
            d = Deal(
                project_name=project_name,
                track=track,
                sub_tag=sub_tag,
                founded_year=founded_year_raw or None,
                title=title,
                team=team,
                round=round_ or "未披露",
                amount=amount,
                investors=investors,
                region=region,
                region_class=region_class,
                detail=detail,
                # 默认值
                valuation="未披露",
                business="",
                importance="mid",
                official_site="",
                verified_date="",
                date_status="in_window",
                date_confidence="high",
                source_url="",
                source_date="",
                sources=[],
                first_seen_window="import_20250625",
            )
            deals.append(d)
        except Exception as e:
            logger.warning(f"[import] 行 {row_idx} ({project_name}) 构造 Deal 失败: {e}")

    return deals


def main():
    p = argparse.ArgumentParser(description="Excel → SQLite 基线导入")
    p.add_argument("--excel", default=DEFAULT_EXCEL, help="Excel 文件路径")
    p.add_argument("--dry-run", action="store_true", help="预览不写入")
    a = p.parse_args()

    xlsx_path = Path(a.excel)
    if not xlsx_path.exists():
        print(f"✗ 文件不存在: {xlsx_path}")
        sys.exit(1)

    print(f"读取: {xlsx_path.name}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    all_deals = []
    track_counts = {}
    for sheet_name in DATA_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  ⚠ 跳过不存在的 sheet: {sheet_name}")
            continue
        ws = wb[sheet_name]
        deals = _import_sheet(ws, sheet_name)
        all_deals.extend(deals)
        track_counts[sheet_name] = len(deals)
        print(f"  [{sheet_name}] {len(deals)} 条")

    wb.close()

    total = len(all_deals)
    print(f"\n总计: {total} 条")
    print(f"分布: {track_counts}")

    if a.dry_run:
        print("\n--- DRY RUN (未写入) ---")
        for d in all_deals[:10]:
            print(f"  [{d.track}] {d.project_name} | {d.round} | {d.amount} | {d.investors[:30] if d.investors else '-'}")
        if total > 10:
            print(f"  ... 还有 {total - 10} 条")
        return

    # 确认
    print(f"\n即将: 清空现有数据库 → 写入 {total} 条")
    print("继续? (y/n): ", end="", flush=True)
    choice = input().strip().lower()
    if choice not in ("y", "yes"):
        print("已取消")
        return

    # 清空 + 写入
    print("清空 deals 表...")
    db.clear_deals()

    print(f"写入 {total} 条...")
    db.upsert(all_deals, window_tag="import_20250625")

    # 验证
    rows = db.all_rows()
    print(f"✓ 数据库现有 {len(rows)} 条记录")

    # 打印各 track 分布验证
    from collections import Counter
    dist = Counter(r["track"] for r in rows)
    for t, c in sorted(dist.items()):
        print(f"  [{t}] {c} 条")


if __name__ == "__main__":
    main()
