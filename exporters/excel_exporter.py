import os
import shutil
from typing import List, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from config.settings import TEMPLATE_PATH
from config.taxonomy import SHEET_BY_TRACK
from models.schema import Deal

_HEADERS = ["项目名称", "细分tag", "成立时间", "标题", "团队", "轮次", "融资金额", "投资方", "业务简介", "地区", "具体信息"]


def _create_template(path: str) -> None:
    """模板缺失时按固定列头 + 5 sheet 自动创建，确保兜底可用。"""
    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    widths = [16, 14, 12, 50, 45, 12, 18, 35, 30, 12, 60]
    for i, name in enumerate(SHEET_BY_TRACK):
        ws = wb.create_sheet(title=name, index=i)
        for ci, h in enumerate(_HEADERS, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font, c.fill, c.alignment, c.border = header_font, header_fill, header_align, border
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + ci)].width = w
        ws.freeze_panes = "A2"
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb.save(path)


def _load_or_create(template: str) -> str:
    if not os.path.exists(template):
        _create_template(template)
    return template


def write_weekly(deals: List[Deal], xlsx_path: str, template: str = TEMPLATE_PATH) -> Tuple[str, List[Deal]]:
    """周报 Excel：剔除 stale（旧闻），严格按模板写入。"""
    keep = [d for d in deals if d.date_status != "stale"]
    shutil.copy(_load_or_create(template), xlsx_path)
    wb = load_workbook(xlsx_path)
    for d in keep:
        sheet_name = SHEET_BY_TRACK.get(d.track)
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        ws.append([
            d.project_name, d.sub_tag, d.founded_year or "",
            d.title, d.team, d.round,
            d.amount or "未披露", d.investors,
            d.business or "", d.region, d.detail,
        ])
    wb.save(xlsx_path)
    return xlsx_path, keep


# 兼容旧引用：从 SQLite 全量重建（非 v4 主路径，仅向后兼容）
def rebuild_excel(template: str = TEMPLATE_PATH, out: str = "output.xlsx") -> str:
    from storage.db import all_rows

    shutil.copy(_load_or_create(template), out)
    wb = load_workbook(out)
    for d in all_rows():
        sheet_name = SHEET_BY_TRACK.get(d["track"])
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        ws.append([
            d["project_name"], d["sub_tag"], d["founded_year"] or "",
            d["title"], d["team"], d["round"],
            d["amount"] or "未披露", d["investors"],
            d.get("business", ""), d["region"], d["detail"],
        ])
    wb.save(out)
    return out
